# reporting/views.py
from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.views.generic import ListView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Count, Q
from django.utils import timezone
from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from datetime import datetime, timedelta
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
import io
import os

from users.models import User
from titres.models import Titre
from demandes.models import Demande
from .models import Report, Dashboard, AuditLog
from .serializers import ReportSerializer, DashboardSerializer, AuditLogSerializer, StatisticsSerializer

class DashboardView(LoginRequiredMixin, ListView):
    template_name = 'reporting/dashboard.html'
    context_object_name = 'stats'
    
    def get_queryset(self):
        return None
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Statistiques générales
        context['total_titres'] = Titre.objects.count()
        context['total_demandes'] = Demande.objects.count()
        context['total_users'] = User.objects.count()
        
        # Statistiques par statut
        context['titres_par_statut'] = dict(
            Titre.objects.values('status').annotate(count=Count('id'))
            .values_list('status', 'count')
        )
        
        context['demandes_par_statut'] = dict(
            Demande.objects.values('status').annotate(count=Count('id'))
            .values_list('status', 'count')
        )
        
        # Titres expirant dans les 30 prochains jours
        date_limite = timezone.now().date() + timedelta(days=30)
        context['titres_expirant'] = Titre.objects.filter(
            date_expiration__lte=date_limite,
            status='approuve'
        ).count()
        
        # Évolution mensuelle
        context['evolution_demandes'] = list(self.get_monthly_evolution())
        
        # Logs récents
        context['recent_logs'] = AuditLog.objects.select_related('user')[:10]
        
        return context
    
    def get_monthly_evolution(self):
        """Calcule l'évolution des demandes sur 6 mois"""
        evolution = []
        for i in range(6):
            date = timezone.now() - timedelta(days=30 * i)
            start_month = date.replace(day=1)
            
            # Calculer le dernier jour du mois
            if start_month.month == 12:
                end_month = start_month.replace(year=start_month.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                end_month = start_month.replace(month=start_month.month + 1, day=1) - timedelta(days=1)
            
            count_demandes = Demande.objects.filter(
                date_soumission__range=[start_month, end_month]
            ).count()
            
            count_titres = Titre.objects.filter(
                date_emission__range=[start_month, end_month]
            ).count()
            
            evolution.append({
                'month': start_month.strftime('%B %Y'),
                'demandes': count_demandes,
                'titres': count_titres
            })
        
        return reversed(evolution)

class ReportViewSet(viewsets.ModelViewSet):
    queryset = Report.objects.all()
    serializer_class = ReportSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        return Report.objects.filter(created_by=self.request.user)
    
    @action(detail=False, methods=['post'])
    def generate_titles_report(self, request):
        """Génère un rapport des titres"""
        filters = request.data.get('filters', {})
        format_type = request.data.get('format', 'pdf')
        
        # Appliquer les filtres
        queryset = Titre.objects.select_related('proprietaire').all()
        if filters.get('status'):
            queryset = queryset.filter(status=filters['status'])
        if filters.get('type'):
            queryset = queryset.filter(type=filters['type'])
        if filters.get('date_debut') and filters.get('date_fin'):
            queryset = queryset.filter(
                date_emission__range=[filters['date_debut'], filters['date_fin']]
            )
        
        # Créer le rapport selon le format
        if format_type == 'pdf':
            return self.generate_pdf_report(queryset, 'Rapport des Titres', 'titres')
        elif format_type == 'excel':
            return self.generate_excel_report(queryset, 'Rapport des Titres', 'titres')
        else:
            return Response({'error': 'Format non supporté'}, 
                          status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['post'])
    def generate_requests_report(self, request):
        """Génère un rapport des demandes"""
        filters = request.data.get('filters', {})
        format_type = request.data.get('format', 'pdf')
        
        queryset = Demande.objects.select_related('demandeur').all()
        if filters.get('status'):
            queryset = queryset.filter(status=filters['status'])
        if filters.get('type_titre'):
            queryset = queryset.filter(type_titre=filters['type_titre'])
        
        if format_type == 'pdf':
            return self.generate_pdf_report(queryset, 'Rapport des Demandes', 'demandes')
        elif format_type == 'excel':
            return self.generate_excel_report(queryset, 'Rapport des Demandes', 'demandes')
    
    def generate_pdf_report(self, queryset, title, report_type='titres'):
        """Génère un rapport PDF"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch)
        styles = getSampleStyleSheet()
        
        # Titre du rapport
        story = []
        title_style = styles['Title']
        title_style.alignment = 1  # Centré
        story.append(Paragraph(title, title_style))
        story.append(Paragraph(f"Généré le: {timezone.now().strftime('%d/%m/%Y à %H:%M')}", styles['Normal']))
        story.append(Spacer(1, 0.2*inch))
        
        # Données du tableau
        if report_type == 'titres':
            data = [['Numéro', 'Type', 'Propriétaire', 'Entreprise', 'Statut', 'Expiration']]
            for titre in queryset:
                data.append([
                    titre.numero_titre or 'N/A',
                    titre.get_type_display(),
                    titre.proprietaire.get_full_name() if titre.proprietaire else 'N/A',
                    titre.entreprise_nom or 'N/A',
                    titre.get_status_display(),
                    titre.date_expiration.strftime('%d/%m/%Y') if titre.date_expiration else 'N/A'
                ])
        else:  # demandes
            data = [['N° Dossier', 'Demandeur', 'Entreprise', 'Type', 'Statut', 'Date']]
            for demande in queryset:
                data.append([
                    demande.numero_dossier or 'En attente',
                    demande.demandeur.get_full_name(),
                    demande.entreprise or 'N/A',
                    demande.get_type_titre_display(),
                    demande.get_status_display(),
                    demande.date_soumission.strftime('%d/%m/%Y')
                ])
        
        # Créer le tableau
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        story.append(table)
        
        # Ajouter les statistiques
        story.append(Spacer(1, 0.3*inch))
        stats_title = Paragraph("Statistiques", styles['Heading2'])
        story.append(stats_title)
        
        total_records = len(queryset)
        stats_text = f"Nombre total d'enregistrements: {total_records}"
        story.append(Paragraph(stats_text, styles['Normal']))
        
        doc.build(story)
        
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{title.replace(" ", "_")}.pdf"'
        
        # Log de l'action
        self.create_audit_log('export', 'Report', title)
        
        return response
    
    def generate_excel_report(self, queryset, title, report_type='titres'):
        """Génère un rapport Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel limite à 31 caractères
        
        # Style pour l'en-tête
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Titre du rapport
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = title
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal="center")
        
        # Date de génération
        ws.merge_cells('A2:F2')
        date_cell = ws['A2']
        date_cell.value = f"Généré le: {timezone.now().strftime('%d/%m/%Y à %H:%M')}"
        date_cell.alignment = Alignment(horizontal="center")
        
        # Ligne vide
        current_row = 4
        
        if report_type == 'titres':
            # En-têtes
            headers = ['Numéro', 'Type', 'Propriétaire', 'Entreprise', 'Statut', 
                      'Date Émission', 'Date Expiration', 'Durée (ans)']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            current_row += 1
            
            # Données
            for titre in queryset:
                row_data = [
                    titre.numero_titre or 'N/A',
                    titre.get_type_display(),
                    titre.proprietaire.get_full_name() if titre.proprietaire else 'N/A',
                    titre.entreprise_nom or 'N/A',
                    titre.get_status_display(),
                    titre.date_emission.strftime('%d/%m/%Y') if titre.date_emission else 'N/A',
                    titre.date_expiration.strftime('%d/%m/%Y') if titre.date_expiration else 'N/A',
                    titre.duree_ans or 'N/A'
                ]
                
                for col, value in enumerate(row_data, 1):
                    ws.cell(row=current_row, column=col, value=value)
                current_row += 1
        
        else:  # demandes
            headers = ['N° Dossier', 'Demandeur', 'Entreprise', 'Email', 'Type', 'Statut', 'Date Soumission']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            current_row += 1
            
            for demande in queryset:
                row_data = [
                    demande.numero_dossier or 'En attente',
                    demande.demandeur.get_full_name(),
                    demande.entreprise or 'N/A',
                    demande.email_contact or 'N/A',
                    demande.get_type_titre_display(),
                    demande.get_status_display(),
                    demande.date_soumission.strftime('%d/%m/%Y')
                ]
                
                for col, value in enumerate(row_data, 1):
                    ws.cell(row=current_row, column=col, value=value)
                current_row += 1
        
        # Ajuster la largeur des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Statistiques
        current_row += 2
        ws.merge_cells(f'A{current_row}:B{current_row}')
        stats_cell = ws[f'A{current_row}']
        stats_cell.value = f"Total d'enregistrements: {len(queryset)}"
        stats_cell.font = Font(bold=True)
        
        # Sauvegarder dans un buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{title.replace(" ", "_")}.xlsx"'
        
        # Log de l'action
        self.create_audit_log('export', 'Report', title)
        
        return response
    
    def create_audit_log(self, action, model_name, description):
        """Créer un log d'audit"""
        try:
            ip_address = self.request.META.get('REMOTE_ADDR')
            user_agent = self.request.META.get('HTTP_USER_AGENT', '')
            
            AuditLog.objects.create(
                user=self.request.user,
                action=action,
                model_name=model_name,
                description=description,
                ip_address=ip_address,
                user_agent=user_agent
            )
        except Exception as e:
            # Ne pas faire échouer la requête à cause du logging
            print(f"Erreur création audit log: {e}")

class DashboardViewSet(viewsets.ModelViewSet):
    serializer_class = DashboardSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        return Dashboard.objects.filter(user=self.request.user)

class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = AuditLog.objects.all()
    serializer_class = AuditLogSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        queryset = AuditLog.objects.select_related('user').all()
        
        # Filtres
        user_id = self.request.query_params.get('user_id')
        action = self.request.query_params.get('action')
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')
        
        if user_id:
            queryset = queryset.filter(user_id=user_id)
        if action:
            queryset = queryset.filter(action=action)
        if date_from:
            queryset = queryset.filter(timestamp__gte=date_from)
        if date_to:
            queryset = queryset.filter(timestamp__lte=date_to)
        
        return queryset

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_statistics(request):
    """API pour récupérer les statistiques du dashboard"""
    try:
        stats = {
            'total_titres': Titre.objects.count(),
            'total_demandes': Demande.objects.count(),
            'total_users': User.objects.count(),
            'titres_actifs': Titre.objects.filter(status='approuve').count(),
            'demandes_en_cours': Demande.objects.filter(status__in=['soumise', 'en_examen']).count(),
            
            # Répartition par type de titre
            'titres_par_type': dict(
                Titre.objects.values('type').annotate(count=Count('id'))
                .values_list('type', 'count')
            ),
            
            # Répartition des demandes par statut
            'demandes_par_statut': dict(
                Demande.objects.values('status').annotate(count=Count('id'))
                .values_list('status', 'count')
            ),
            
            # Évolution mensuelle
            'evolution_mensuelle': list(get_monthly_stats()),
            
            # Titres expirant bientôt
            'titres_expirant_30j': Titre.objects.filter(
                date_expiration__lte=timezone.now().date() + timedelta(days=30),
                status='approuve'
            ).count(),
        }
        
        serializer = StatisticsSerializer(stats)
        return Response(serializer.data)
        
    except Exception as e:
        return Response(
            {'error': f'Erreur lors de la récupération des statistiques: {str(e)}'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

def get_monthly_stats():
    """Statistiques mensuelles pour les graphiques"""
    for i in range(6):
        date = timezone.now() - timedelta(days=30 * i)
        start_month = date.replace(day=1)
        
        # Calculer le dernier jour du mois
        if start_month.month == 12:
            end_month = start_month.replace(year=start_month.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            end_month = start_month.replace(month=start_month.month + 1, day=1) - timedelta(days=1)
        
        titres_count = Titre.objects.filter(
            date_emission__range=[start_month, end_month]
        ).count()
        
        demandes_count = Demande.objects.filter(
            date_soumission__range=[start_month, end_month]
        ).count()
        
        yield {
            'month': start_month.strftime('%m/%Y'),
            'titres': titres_count,
            'demandes': demandes_count
        }
      